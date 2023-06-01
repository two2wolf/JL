from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service
import datetime
import time
import re

# 替换为你的WebDriver路径
webdriver_path = "C:/Users/tms/Downloads/chromedriver_win32/chromedriver.exe"

# 使用Chrome WebDriver
service = Service(executable_path=webdriver_path)
driver = webdriver.Chrome(service=service)

wait = WebDriverWait(driver, 10)

import openpyxl
from openpyxl import load_workbook

def read_excel_worksheet(file_path, sheet_name):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表
    worksheet = workbook[sheet_name]

    # 读取工作表内容
    content = []
    for row in worksheet.iter_rows():
        row_data = [cell.value for cell in row]
        content.append(row_data)

    return content

if __name__ == "__main__":
    file_path = "D:\\360MoveData\\Users\\tms\\Desktop\\查处偷逃费车辆处理单-空工作簿.xlsx"  # 用实际的文件路径替换
    sheet_name = "Sheet1"  # 用实际的工作表名称替换

    content = read_excel_worksheet(file_path, sheet_name)

    for row in content:
        print(row)

content = read_excel_worksheet(file_path, sheet_name)
title = content[0][0]
toll_station = content[1][1]
check_time = content[2][1]
license_plate = content[3][1]
card_number = content[4][1]
plate_color = content[5][1]
vehicle_color = content[6][1]
entry_type = content[7][1]
exit_type = content[8][1]
entry_station = content[9][1]
entry_time = content[10][1]
evasion_method = content[11][1]
discovery_method = content[12][1]
check_process = content[13][1]
check_unit = content[14][1]
pre_check_fee = content[15][1]
post_check_fee = content[16][1]
recovered_fee = content[17][1]
main_investigator = [content[18][1], content[18][2]]
main_investigator_1 = content[18][1]
main_investigator_2 = content[18][2]
squad_investigator = [content[19][1], content[19][2]]
squad_investigator_1 = content[19][1]
squad_investigator_2 = content[19][2]
external_investigator = content[20][1]
reward_distribution = content[21][1]
reward_main_investigator = content[22][1]
reward_main_investigator_2 = content[23][1]
reward_squad_investigator = content[24][1]
reward_squad_investigator_2 = content[25][1]
reward_external_investigator = content[26][1]
total_reward = content[27][1]
filler = content[28][1]

check_year, check_month, check_day, hour, minute = check_time.year, check_time.month, check_time.day, check_time.hour, check_time.minute

check_month_str = str(int(check_month))
check_day_str = str(int(check_day))

# 访问目标网站
driver.get("http://124.128.225.22:3536/jlsfjs/a/login")


# 定位输入框并输入数据
def input_text(element_name, text):
    input_element = driver.find_element(By.NAME, element_name)
    input_element.send_keys(text)

input_text("username_", "jlguohongyun")
input_text("password_", "Sfzh@2020!")

# 定位并单击提交按钮
submit_button = driver.find_element(By.CSS_SELECTOR, 'input.btn.btn-large.btn-primary')
submit_button.click()

# 访问目标网站
driver.get("http://124.128.225.22:3536/jlsfjs/a/sfjs/sfjcjc/jcescapeveh/form?check=add&pageNo=1&pageSize=15###")

# 定位元素
element1 = driver.find_element(By.ID, "stationIdName")
driver.execute_script("arguments[0].click();", element1)

wait = WebDriverWait(driver, 30)

# 定位到包含iframe的div元素
jbox_element = wait.until(EC.presence_of_element_located((By.ID, 'jbox')))

# 定位到iframe元素
iframe_element = wait.until(EC.presence_of_element_located((By.ID, 'jbox-iframe')))

# 切换到iframe
driver.switch_to.frame(iframe_element)

# 在iframe中定位到<span>元素
target_span = wait.until(EC.presence_of_element_located((By.ID, 'tree_1_span')))

# 使用JavaScript点击<span>元素
driver.execute_script("arguments[0].click();", target_span)
print(target_span.is_displayed())  # 应返回True，如果返回False说明元素不可见
print(target_span.is_enabled())    # 应返回True，如果返回False说明元素不可用
# 创建 ActionChains 对象
actions = ActionChains(driver)

# 对目标元素执行双击操作
actions.double_click(target_span).perform()

# 定位到<license_plate>元素
input_license_plate_element = wait.until(EC.presence_of_element_located((By.ID, 'vehNumber')))

# 输入内容license_plate
input_license_plate_element.send_keys(license_plate)

# 定位到 <card_number> 元素
input_card_number_element = wait.until(EC.presence_of_element_located((By.NAME, 'cardNumber')))

# 输入内容card_number
input_card_number_element.send_keys(card_number)

# 定位到<entry_station>元素
input_entry_station_element = wait.until(EC.presence_of_element_located((By.ID, 'enstationName')))

# 输入内容entry_station
input_entry_station_element.send_keys(entry_station)

# 定位到<vehicle_color>元素
input_vehicle_color_element = wait.until(EC.presence_of_element_located((By.ID, 'vehColor')))

# 输入内容vehicle_color
input_vehicle_color_element.send_keys(vehicle_color)

# 定位到<check_process>元素
input_check_process_element = wait.until(EC.presence_of_element_located((By.ID, 'jcProcedure')))

# 输入内容vehicle_color
input_check_process_element.send_keys(check_process)

# 定位元素 s2id_entranceType
element_s2id_entranceType = driver.find_element(By.ID, 's2id_entranceType')

# 定位元素 select2-choice0
element_select2_choice0 = element_s2id_entranceType.find_element(By.CLASS_NAME, 'select2-choice')

# 定位元素 span_entry_type
element_span_entry_type = element_select2_choice0.find_element(By.CLASS_NAME, 'select2-chosen')

# 点击元素 span_entry_type
element_span_entry_type.click()


# 等待下拉菜单的选项出现
entranceType_list = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//ul[contains(@class, "select2-results")]/li')))

# 找到与entry_type值匹配的选项并点击
for option in entranceType_list:
    if option.text == entry_type:
        option.click()
        break


# 根据id属性值定位元素s2id_vlpcolor
element_s2id_vlpcolor = wait.until(EC.presence_of_element_located((By.ID, 's2id_vlpcolor')))

# 在元素s2id_vlpcolor内，根据类名定位元素select2-choice1
element_select2_choice1 = element_s2id_vlpcolor.find_element(By.CLASS_NAME, 'select2-choice')

# 在元素select2-choice1内，根据类名定位元素span_plate_color
element_span_plate_color = element_select2_choice1.find_element(By.CLASS_NAME, 'select2-chosen')

# 点击元素span_plate_color
element_span_plate_color.click()

# 等待下拉菜单中的select2-results元素出现
vlpcolor_list = wait.until(EC.presence_of_all_elements_located((By.XPATH, '//ul[contains(@class, "select2-results")]/li')))


# 找到与plate_color值匹配的选项并点击
for option in vlpcolor_list:
    if option.text == plate_color:
        option.click()
        break

# 根据id属性值定位元素s2id_exitType
s2id_exitType = wait.until(EC.presence_of_element_located((By.ID, 's2id_exitType')))

# 在s2id_exitType内，根据类名定位select2_choice2
select2_choice2 = s2id_exitType.find_element(By.CLASS_NAME, 'select2-choice')

# 在select2_choice2内，根据类名定位span_exitType_element
span_exitType_element = select2_choice2.find_element(By.CLASS_NAME, 'select2-chosen')

# 点击
span_exitType_element.click()

# 等待下拉菜单中的选项出现
options_list = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.select2-results li')))

# 找到与exitType值匹配的选项并点击
for option in options_list:
    if option.text == exit_type:
        option.click()
        break
# 定位包含所有checkbox的元素
checkboxes = wait.until(EC.presence_of_all_elements_located((By.NAME, 'dealTypeArray')))

# 对于每一个checkbox，获取它的父元素（label元素）的文本，如果文本和discovery_method值匹配，那么就点击这个checkbox
for checkbox in checkboxes:
    label = checkbox.find_element(By.XPATH, './..')
    label_text = label.text.strip()
    if label_text == discovery_method:
        checkbox.click()
        break
# 定位到元素
dealUnit1Name_element = wait.until(EC.presence_of_element_located((By.ID, 'dealUnit1Name')))
dealUnit2Name_element = wait.until(EC.presence_of_element_located((By.ID, 'dealUnit2Name')))
# 根据discovery_method的值选择点击哪个元素
if discovery_method == '现场发现查处':
    dealUnit1Name_element.click()
elif discovery_method == '灰（黑）名单拦截':
    dealUnit2Name_element.click()

# 定位到<dealUnit1Namediv>元素
try:
    dealUnit1Name_div_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "jbox")))
    dealUnit1Name_iframe_element = dealUnit1Name_div_element.find_element(By.ID, "jbox-iframe")
except:
    print("Element not found")

# 在<dealUnit1Namediv>元素内部定位到<iframe>
dealUnit1Name_iframe_element = dealUnit1Name_div_element.find_element(By.ID, "jbox-iframe")

# 切换到<iframe>
driver.switch_to.frame(dealUnit1Name_iframe_element)

# 在<iframe>内部找到目标元素
try:
    # 等待目标元素加载完成
    target_element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "tree_210_a"))
    )
    # 双击目标元素
    ActionChains(driver).double_click(target_element).perform()
except:
    print("无法定位到元素。")
    
# 切换回主文档
driver.switch_to.default_content()

# 定位 input 元素 by id
beforeMoney_element = driver.find_element(By.ID, 'beforeMoney')
afterMoney_element = driver.find_element(By.ID, 'afterMoney')


# 输入预检查费用和检查后费用
beforeMoney_element.send_keys(pre_check_fee)
afterMoney_element.send_keys(post_check_fee)

# 第一步，定位到外层tr元素
majorJcerArray_tr = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//td[@class="tit"]/..')))
majorJcerArray_div = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '[id="s2id_majorJcerArray"]')))
majorJcerArray_ul = majorJcerArray_div.find_element(By.CLASS_NAME, 'select2-choices')


# 点击ul元素以打开下拉菜单
majorJcerArray_ul.click()

try:
    # Wait up to 30 seconds for the element to become available
    majorJcerArray_ul_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "select2-drop")))
except TimeoutException:
    print("Timed out waiting for element to appear")

if main_investigator_1:
    # 获取所有的下拉列表项
    items = majorJcerArray_ul_element.find_elements(By.CLASS_NAME, 'select2-result-label')
    # 遍历所有列表项
    for item in items:
        # 如果列表项的文本与main_investigator_1匹配，点击它
        if main_investigator_1 in item.text:
            ActionChains(driver).move_to_element(item).click(item).perform()
            break
else:
# 如果main_investigator_2不存在或为空，按键盘上的esc键，取消操作
   ActionChains(driver).send_keys(Keys.ESCAPE).perform()

# 第一步，定位到外层tr元素
majorJcerArray_tr = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//td[@class="tit"]/..')))
majorJcerArray_div = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.CSS_SELECTOR, '[id="s2id_majorJcerArray"]')))
majorJcerArray_ul = majorJcerArray_div.find_element(By.CLASS_NAME, 'select2-choices')


# 点击ul元素以打开下拉菜单
majorJcerArray_ul.click()

try:
    # Wait up to 30 seconds for the element to become available
    majorJcerArray_ul_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "select2-drop")))
except TimeoutException:
    print("Timed out waiting for element to appear")

if main_investigator_2:
    # 获取所有的下拉列表项
    items = majorJcerArray_ul_element.find_elements(By.CLASS_NAME, 'select2-result-label')
    # 遍历所有列表项
    for item in items:
        # 如果列表项的文本与main_investigator_1匹配，点击它
        if main_investigator_2 in item.text:
            ActionChains(driver).move_to_element(item).click(item).perform()
            break
else:
    # 如果main_investigator_2不存在或为空，按键盘上的esc键，取消操作
    ActionChains(driver).send_keys(Keys.ESCAPE).perform()
# 定位到外层tr元素
squad_investigator_tr = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//td[text()="中队协办人员"]/..')))

# 定位到td元素
squad_investigator_td = squad_investigator_tr.find_element(By.CSS_SELECTOR, 'td[colspan="4"]')

# 定位到div元素
squad_investigator_div = squad_investigator_td.find_element(By.ID, 's2id_middlePersonArray')

# 定位到ul元素
squad_investigator_ul = squad_investigator_div.find_element(By.CLASS_NAME, 'select2-choices')

# 点击ul元素以打开下拉菜单
squad_investigator_ul.click()

try:
    # Wait up to 30 seconds for the element to become available
    squad_investigator_ul_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "select2-drop")))
except TimeoutException:
    print("Timed out waiting for element to appear")

if squad_investigator_1:
    # 获取所有的下拉列表项
    items = squad_investigator_ul_element.find_elements(By.CLASS_NAME, 'select2-result-label')
    # 遍历所有列表项
    for item in items:
        # 如果列表项的文本与squad_investigator_1匹配，点击它
        if squad_investigator_1 in item.text:
            ActionChains(driver).move_to_element(item).click(item).perform()
            break
else:
    # 如果squad_investigator_1不存在或为空，按键盘上的esc键，取消操作
    ActionChains(driver).send_keys(Keys.ESCAPE).perform()

# 定位到外层tr元素
squad_investigator_tr = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//td[text()="中队协办人员"]/..')))

# 定位到td元素
squad_investigator_td = squad_investigator_tr.find_element(By.CSS_SELECTOR, 'td[colspan="4"]')

# 定位到div元素
squad_investigator_div = squad_investigator_td.find_element(By.ID, 's2id_middlePersonArray')

# 定位到ul元素
squad_investigator_ul = squad_investigator_div.find_element(By.CLASS_NAME, 'select2-choices')

# 点击ul元素以打开下拉菜单
squad_investigator_ul.click()

try:
    # Wait up to 30 seconds for the element to become available
    squad_investigator_ul_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "select2-drop")))
except TimeoutException:
    print("Timed out waiting for element to appear")

if squad_investigator_2:
    # 获取所有的下拉列表项
    items = squad_investigator_ul_element.find_elements(By.CLASS_NAME, 'select2-result-label')
    # 遍历所有列表项
    for item in items:
        # 如果列表项的文本与squad_investigator_1匹配，点击它
        if squad_investigator_2 in item.text:
            ActionChains(driver).move_to_element(item).click(item).perform()
            break
else:
    # 如果squad_investigator_1不存在或为空，按键盘上的esc键，取消操作
    ActionChains(driver).send_keys(Keys.ESCAPE).perform()

# 定位到外层tr元素
external_investigator_tr = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, '//td[text()="本单位或外部协办人员"]/..')))

# 定位到td元素
external_investigator_td = external_investigator_tr.find_element(By.CSS_SELECTOR, 'td[colspan="4"]')

# 定位到div元素
external_investigator_div = external_investigator_td.find_element(By.ID, 's2id_assistPersonArray')

# 定位到ul元素
external_investigator_ul = external_investigator_div.find_element(By.CLASS_NAME, 'select2-choices')

# 点击ul元素以打开下拉菜单
external_investigator_ul.click()

try:
    # Wait up to 30 seconds for the element to become available
    external_investigator_ul_element = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "select2-drop")))
except TimeoutException:
    print("Timed out waiting for element to appear")

if external_investigator:
    # 获取所有的下拉列表项
    items = external_investigator_ul_element.find_elements(By.CLASS_NAME, 'select2-result-label')
    # 遍历所有列表项
    for item in items:
        # 如果列表项的文本与external_investigator匹配，点击它
        if external_investigator in item.text:
            ActionChains(driver).move_to_element(item).click(item).perform()
            break
else:
    # 如果external_investigator不存在或为空，按键盘上的esc键，取消操作
    ActionChains(driver).send_keys(Keys.ESCAPE).perform()
    
# 切换回主文档
driver.switch_to.default_content()

# 定位table_form_element
table_form_element = driver.find_element(By.CSS_SELECTOR, ".table-form")
tbody_element = table_form_element.find_element(By.XPATH,'//tbody')

jcDate_tr_element = tbody_element.find_element(By.XPATH,'//tr[.//input[@id="jcDate"]]')

jcDate_td_element = jcDate_tr_element.find_element(By.XPATH,'//td[.//input[@id="jcDate"]]')

jcDate_input_element = jcDate_tr_element.find_element(By.ID, "jcDate")

driver.execute_script("arguments[0].click();", jcDate_input_element)

# 等待10秒，让页面加载完成
time.sleep(10)


# 寻找iframe元素
jcDate_iframe_element = driver.find_element(By.CSS_SELECTOR, 'iframe[style="width: 202px; height: 245px;"]')

# 切换到iframe
driver.switch_to.frame(jcDate_iframe_element)
# 在iframe中找到元素
WdateDiv_element = driver.find_element(By.CLASS_NAME, 'WdateDiv')

# 定位dpTitle
dpTitle_div_element = WdateDiv_element.find_element(By.ID, 'dpTitle')
dpTime_div_element = WdateDiv_element.find_element(By.ID, 'dpTime')
# 定位这个input元素
input_month_elements = dpTitle_div_element.find_elements(By.CLASS_NAME, 'yminput')

actions = ActionChains(driver)

# 对第一个input元素的操作
actions.move_to_element(input_month_elements[0])  # 移动到第一个input元素
actions.click()  # 模拟鼠标点击
actions.send_keys(Keys.BACK_SPACE)  # 清空元素中的内容
actions.send_keys(check_month_str)  # 在元素中输入文本

actions.perform()  # 执行以上操作

# 清空之前的actions
actions.reset_actions()

# 对第二个input元素的操作
actions.move_to_element(input_month_elements[1])  # 移动到第二个input元素
actions.click()  # 模拟鼠标点击
actions.send_keys(Keys.BACK_SPACE)  # 清空元素中的内容
actions.send_keys(check_year)  # 在元素中输入文本

actions.perform()  # 执行以上操作

# 定位"时"的输入框
hour1_input = dpTime_div_element.find_element(By.CSS_SELECTOR, "input.tB")

# 定位"分"的输入框
minute1_input = dpTime_div_element.find_element(By.CSS_SELECTOR, "input.tE")

# 定位"秒"的输入框
second1_input = dpTime_div_element.find_element(By.CSS_SELECTOR, "input.tE[disabled]")

# 在"时"的输入框中输入小时
actions.move_to_element(hour1_input).click().send_keys(hour).perform()

# 在"分"的输入框中输入分钟
actions.move_to_element(minute1_input).click().send_keys(minute).perform()

ActionChains(driver).send_keys(Keys.ESCAPE).perform()

WdayTable_element = WdateDiv_element.find_element(By.CSS_SELECTOR,"table.WdayTable")

WdayTable_element = WdateDiv_element.find_element(By.CSS_SELECTOR,"table.WdayTable")

# Find the tbody element within the table
WdayTbody_element = WdayTable_element.find_element(By.TAG_NAME, 'tbody')

# Find all td elements within the tbody with class Wday, Wselday or Wwday
day_elements = WdayTbody_element.find_elements(By.XPATH, ".//td[contains(@class, 'Wday') or contains(@class, 'Wselday') or contains(@class, 'Wwday')]")
# Using WebDriverWait to wait for the date element to appear and be clickable
date_element = WebDriverWait(WdayTbody_element, 10).until(EC.element_to_be_clickable((By.XPATH, f"//td[starts-with(@onclick, 'day_Click({check_year},{check_month_str},{check_day_str})')]")))
# Using JavaScript to perform click action
driver.execute_script("arguments[0].click();", date_element)
    

# 等待10秒，让页面加载完成

time.sleep(10)
